#!/usr/bin/env python

from openpyxl import load_workbook
import sys
import os
import json



def check_path(filepath):
	if os.path.exists('./runtime/'+filepath):
		return True
	sys.exit('[!] filepath:%s not exists'%filepath)

wb = load_workbook("res.xlsx")
print(wb.sheetnames)
sheet_lee = wb.get_sheet_by_name("1")
sheet_ins = wb.get_sheet_by_name("2")
sheet_haozi = wb.get_sheet_by_name("3")

sheets = [sheet_lee,sheet_ins,sheet_haozi]

res = {}

for sheet in sheets:
	filename = sheet['A']
	begin = sheet['B']
	end = sheet['C']

	# the length of all columns should be the same
	assert len(filename)==len(begin)
	assert len(begin)==len(end)

	for i in range(len(filename)):
            if res.has_key(filename[i].value):
                res[filename[i].value].append({"begin":begin[i],"end":end[i]})
            else:
                res[filename[i].value] = []




print res
open('submit.txt','w').write(json.dumps(res))







